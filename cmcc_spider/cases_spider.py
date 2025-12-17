import requests
from lxml import etree
from concurrent.futures import ThreadPoolExecutor, as_completed
from openpyxl import Workbook, load_workbook
import os
import time

# ================== 基础配置 ==================

BASE_URL = "http://www.cmcc-dlut.cn"
EXCEL_FILE = "cases.xlsx"

HEADERS = {
    "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36",
    "Referer": BASE_URL
}

MAX_WORKERS = 2          # 并发刻意压低
PAGE_SLEEP = 5           # 每页列表休息
DETAIL_SLEEP = 1         # 每条详情休息
ERROR_SLEEP = 300        # 500 / 异常后休息 5 分钟

# ================== 字段映射 ==================

FIELD_MAP = {
    # 文本字段
    '案例编号：': 'case_id',
    '被浏览次数：': 'case_view_count',
    '案例名称：': 'case_title',
    '译名：': 'case_translated_title',
    '案例作者：': 'case_author',
    '作者单位：': 'case_author_affiliation',
    '指导者：': 'case_advisor',
    '译者：': 'case_translator',
    '案例企业名称：': 'case_company_name',
    '行业：': 'case_industry',
    '规模：': 'case_company_scale',
    '案例涉及的职能领域：': 'case_functional_area',
    '案例语种：': 'case_language',
    '案例正文页数（页）：': 'case_page_count',
    '案例类型：': 'case_type',
    '中文关键词：': 'keywords_zh',
    '英文关键词：': 'keywords_en',
    '中文摘要：': 'abstract_zh',
    '英文摘要：': 'abstract_en',
    '适用对象：': 'target_audience',
    '编写方式：': 'writing_method',
    '案例年代：': 'case_year',
    '适用课程：': 'applicable_courses',
    '案例涉及理论知识：': 'theoretical_knowledge',
    '案例入库时间：': 'case_created_time',

    # 链接字段
    '查看全文:': 'full_text_url',
    '查看案例使用说明': 'usage_guide_url',
    '查看微案例英文正文': 'micro_case_en_full_text_url',
    '查看微案例英文使用说明': 'micro_case_en_guide_url',
}

# ================== 网络请求 ==================

def fetch(url, retry=2):
    for i in range(retry):
        try:
            r = requests.get(url, headers=HEADERS, timeout=10)
            r.raise_for_status()
            return r.text
        except Exception:
            time.sleep(2 + i)
    return None

# ================== 字段解析并提取 ==================

def get_text_value(ele, label):
    v = ele.xpath(
        f".//th[contains(text(),'{label}')]/following-sibling::td[1]/text()"
    )
    return v[0].strip() if v else None

def get_link_value(ele, label):
    v = ele.xpath(
        f".//th[contains(text(),'{label}')]/following-sibling::td[1]/a/@href"
    )
    return BASE_URL + v[0] if v else None

def extract_case_data(content):
    data = {}
    for label, field in FIELD_MAP.items():
        if field.endswith("_url"):
            data[field] = get_link_value(content, label)
        else:
            data[field] = get_text_value(content, label)
    return data

# ================== 详情页 ==================

def parse_detail(url):
    html = fetch(url)
    if not html:
        return None

    tree = etree.HTML(html)
    content = tree.xpath("//div[@class='content_box']")
    if not content:
        return None

    data = extract_case_data(content[0])
    time.sleep(DETAIL_SLEEP)
    return data

# 对数据有缺失的页面进行重爬
def re_crawl_pages(pages):
    print(f"\n 开始补爬页面: {pages}")
    all_rows = []

    for page in pages:
        list_url = f"{BASE_URL}/Cases/Latest/{page}"
        print(f" 补爬第 {page} 页")

        html = fetch(list_url)
        if not html:
            print(f" 第 {page} 页列表失败，跳过")
            continue

        tree = etree.HTML(html)
        links = tree.xpath('//div[@class="result1"]//td[@class="a_left"]/a/@href')

        if not links:
            print(f" 第 {page} 页无数据")
            continue

        for l in links:
            detail_url = BASE_URL + l
            data = parse_detail(detail_url)
            if data:
                all_rows.append(data)
            time.sleep(2)

        time.sleep(10)  # 每页补爬完强制休息

    write_to_excel(all_rows)


# ================== Excel 写入 ==================

def write_to_excel(rows):
    if not rows:
        return

    if os.path.exists(EXCEL_FILE):
        wb = load_workbook(EXCEL_FILE)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.append(list(FIELD_MAP.values()))

    for row in rows:
        ws.append([row.get(col) for col in FIELD_MAP.values()])

    wb.save(EXCEL_FILE)
    print(f" 写入 Excel {len(rows)} 条")

# 对重爬的页面数据进行去重
def deduplicate_excel():
    if not os.path.exists('cases.xlsx'):
        print(" Excel 不存在，无法去重")
        return

    wb = load_workbook(EXCEL_FILE)
    ws = wb.active

    rows = list(ws.iter_rows(values_only=True))
    header = rows[0]
    data_rows = rows[1:]

    case_id_index = header.index("case_id")

    seen = set()
    unique_rows = []

    for row in data_rows:
        case_id = row[case_id_index]
        if case_id and case_id not in seen:
            seen.add(case_id)
            unique_rows.append(row)

    # 重写 Excel
    new_wb = Workbook()
    new_ws = new_wb.active
    new_ws.append(header)

    for r in unique_rows:
        new_ws.append(r)

    new_wb.save(EXCEL_FILE)
    print(f" 去重完成，保留 {len(unique_rows)} 条唯一案例")


# ================== 主爬虫逻辑 ==================

def crawl(start_page, end_page):
    page = start_page

    while page <= end_page:
        list_url = f"{BASE_URL}/Cases/Latest/{page}"
        print(f"\n 正在爬第 {page} 页")

        html = fetch(list_url)
        if not html:
            print(" 列表页失败，休息 5 分钟")
            time.sleep(ERROR_SLEEP)
            continue

        tree = etree.HTML(html)
        links = tree.xpath('//div[@class="result1"]//td[@class="a_left"]/a/@href')

        if not links:
            print(" 当前页无数据，停止")
            break

        detail_urls = [BASE_URL + l for l in links]
        results = []

        with ThreadPoolExecutor(max_workers=MAX_WORKERS) as pool:
            futures = [pool.submit(parse_detail, u) for u in detail_urls]
            for f in as_completed(futures):
                r = f.result()
                if r:
                    results.append(r)

        write_to_excel(results)

        page += 1
        time.sleep(PAGE_SLEEP)

    print(" 爬取完成")


if __name__ == "__main__":
    # 爬取全部页面内的数据
    crawl(start_page=1, end_page=569)
    #TODO 爬虫完成后部分页面可能出现数据缺失，在下面输入对应的页数进行重新爬取
    # 1. 补爬 172、173 页,并对excel内数据去重
    # re_crawl_pages([172, 173])
    # deduplicate_excel()
